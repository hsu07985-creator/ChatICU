#!/usr/bin/env python3
"""Build the canonical drug formulary CSV from multiple sources.

Primary source:  陽明院區常備藥品明細表1150401.xlsx  (1,568 codes, ATC 99.9%)
Secondary:       陽明抗生素清單 20260303.xlsx        (122 codes, is_antibiotic flag)
Tertiary:        FHIR功能/藥物標準化/atc_drugs.csv    (kidney_relevant + rxnorm_cui for overlapping codes)

Output:          backend/app/fhir/code_maps/drug_formulary.csv
                 backend/app/fhir/code_maps/drug_formulary_gaps.csv
                   (DB ODR_CODEs with no formulary entry — for manual curation)

Run whenever the formulary is updated:
    python3 backend/scripts/build_formulary_csv.py
"""
from __future__ import annotations

import asyncio
import csv
import os
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

BACKEND_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = BACKEND_ROOT.parent
sys.path.insert(0, str(BACKEND_ROOT))

FORMULARY_XLSX = REPO_ROOT / "陽明院區常備藥品明細表1150401.xlsx"
ABX_XLSX = REPO_ROOT / "陽明抗生素清單 20260303.xlsx"
ATC_DRUGS_CSV = REPO_ROOT / "FHIR功能" / "藥物標準化" / "atc_drugs.csv"

OUTPUT_DIR = BACKEND_ROOT / "app" / "fhir" / "code_maps"
FORMULARY_CSV = OUTPUT_DIR / "drug_formulary.csv"
GAPS_CSV = OUTPUT_DIR / "drug_formulary_gaps.csv"


def load_formulary_xlsx() -> dict[str, dict[str, Any]]:
    """Parse yangming formulary xlsx → {ODR_CODE: row}.

    Formulary columns: 處置代碼 | 成份及含量 | 單位 | 商品名 | atc code
    Duplicates (18 codes) are collapsed; first wins. Same-code duplicates are
    usually different pack sizes of the same drug — ATC is invariant.
    """
    import openpyxl

    wb = openpyxl.load_workbook(FORMULARY_XLSX, data_only=True, read_only=True)
    ws = wb["工作表1"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header

    out: dict[str, dict[str, Any]] = {}
    for code, ingredient, unit, brand, atc in rows:
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue
        # Dedup: first row wins
        if code in out:
            continue
        out[code] = {
            "odr_code": code,
            "ingredient": (ingredient or "").strip() if ingredient else "",
            "unit": (unit or "").strip() if unit else "",
            "brand_name": (brand or "").strip() if brand else "",
            "atc_code": (atc or "").strip() if atc else "",
        }
    wb.close()
    return out


def load_abx_xlsx() -> dict[str, dict[str, Any]]:
    """Parse antibiotic list xlsx → {ODR_CODE: row}.

    Columns: 醫令代碼 | 醫令名稱 | 醫令名稱2(中文) | 劑型
    """
    import openpyxl

    wb = openpyxl.load_workbook(ABX_XLSX, data_only=True, read_only=True)
    ws = wb["工作表1"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    out: dict[str, dict[str, Any]] = {}
    for code, name, name_zh, form in rows:
        if not code:
            continue
        code = str(code).strip()
        if not code:
            continue
        out[code] = {
            "odr_code": code,
            "name": (name or "").strip() if name else "",
            "name_zh": (name_zh or "").strip() if name_zh else "",
            "form": (form or "").strip() if form else "",
        }
    wb.close()
    return out


def load_atc_drugs_csv() -> dict[str, dict[str, Any]]:
    """Parse the legacy CKD-focused atc_drugs.csv for supplementary fields.

    Columns: odr_code,drug_name,generic_name,atc_code,rxnorm_cui,kidney_relevant,notes
    """
    out: dict[str, dict[str, Any]] = {}
    if not ATC_DRUGS_CSV.exists():
        return out
    with ATC_DRUGS_CSV.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            code = row.get("odr_code", "").strip()
            if code:
                out[code] = row
    return out


def get_db_odr_frequencies() -> dict[str, tuple[int, str]]:
    """Query Supabase for distinct ODR_CODEs in medications with frequency + sample name.

    Returns: {ODR_CODE: (row_count, sample_name)}
    """
    import asyncio

    from sqlalchemy import text
    from sqlalchemy.ext.asyncio import create_async_engine

    def _db_url() -> str | None:
        env_path = os.environ.get("SYNC_ENV_PATH")
        env_file = Path(env_path) if env_path else BACKEND_ROOT / ".env.his-sync"
        if not env_file.exists():
            return None
        for line in env_file.read_text().splitlines():
            line = line.strip()
            if line.startswith("DATABASE_URL="):
                url = line.split("=", 1)[1].strip().strip('"').strip("'")
                if url.startswith("postgresql://"):
                    url = url.replace("postgresql://", "postgresql+asyncpg://", 1)
                return url
        return None

    url = _db_url()
    if not url:
        print("[WARN] No DATABASE_URL; skipping DB gap analysis.", file=sys.stderr)
        return {}

    async def _run() -> dict[str, tuple[int, str]]:
        engine = create_async_engine(
            url,
            connect_args={
                "prepared_statement_cache_size": 0,
                "statement_cache_size": 0,
            },
        )
        out: dict[str, tuple[int, str]] = {}
        async with engine.connect() as conn:
            r = await conn.execute(text(
                "SELECT order_code, COUNT(*), MAX(name) "
                "FROM medications "
                "WHERE order_code IS NOT NULL AND order_code <> '' "
                "GROUP BY order_code"
            ))
            for code, cnt, name in r:
                out[code] = (cnt, name or "")
        await engine.dispose()
        return out

    return asyncio.run(_run())


def load_gap_fills() -> dict[str, dict[str, str]]:
    """Load hand-curated ATC values previously written into drug_formulary_gaps.csv.

    Only rows with a non-empty `suggested_atc` are picked up. This lets the
    operator fill in gaps in a plain CSV and re-run the build to merge them
    into the main formulary CSV.
    """
    if not GAPS_CSV.exists():
        return {}
    out: dict[str, dict[str, str]] = {}
    with GAPS_CSV.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            code = (row.get("odr_code") or "").strip()
            atc = (row.get("suggested_atc") or "").strip()
            if code and atc:
                out[code] = {
                    "atc_code": atc,
                    "sample_name": (row.get("sample_name") or "").strip(),
                    "notes": (row.get("notes") or "").strip(),
                }
    return out


def merge_and_write(
    formulary: dict[str, dict[str, Any]],
    abx: dict[str, dict[str, Any]],
    legacy: dict[str, dict[str, Any]],
    db_freq: dict[str, tuple[int, str]],
    gap_fills: dict[str, dict[str, str]],
) -> tuple[int, int, int]:
    """Merge into a unified formulary CSV.

    Output schema:
        odr_code, atc_code, ingredient, brand_name, unit,
        is_antibiotic, kidney_relevant, rxnorm_cui, source, notes

    `source` tracks which input(s) supplied the code:
        - formulary       : in yangming main formulary (has ATC)
        - formulary+abx   : in both
        - abx_only        : in ABX list only (needs ATC curation)
        - legacy_only     : only in old atc_drugs.csv (CKD-focused, may be noise)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    all_codes = set(formulary) | set(abx) | set(legacy) | set(gap_fills)
    rows: list[dict[str, Any]] = []

    for code in sorted(all_codes):
        f = formulary.get(code)
        a = abx.get(code)
        l = legacy.get(code)

        g = gap_fills.get(code)
        if f and a:
            source = "formulary+abx"
        elif f:
            source = "formulary"
        elif a and g:
            source = "abx+manual"
        elif a:
            source = "abx_only"
        elif g:
            source = "manual"
        else:
            source = "legacy_only"

        # ATC precedence: formulary > manual gap fill > legacy
        atc = (
            (f.get("atc_code") if f else "")
            or (g.get("atc_code", "") if g else "")
            or (l.get("atc_code", "").strip() if l else "")
        )

        # Ingredient / brand: formulary > abx.name > legacy.drug_name
        ingredient = (
            (f.get("ingredient") if f else "")
            or (a.get("name") if a else "")
            or (g.get("sample_name", "") if g else "")
            or (l.get("generic_name", "") if l else "")
        )
        brand = (
            (f.get("brand_name") if f else "")
            or (a.get("name_zh") if a else "")
            or (l.get("drug_name", "") if l else "")
        )
        unit = (f.get("unit") if f else "") or ""

        # rxnorm_cui / kidney_relevant / notes come from legacy only
        rxnorm_cui = (l.get("rxnorm_cui", "").strip() if l else "")
        kidney_rel_raw = (l.get("kidney_relevant", "").strip() if l else "")
        kidney_relevant = ""
        if kidney_rel_raw in {"1", "0"}:
            kidney_relevant = kidney_rel_raw
        notes = (l.get("notes", "").strip() if l else "")

        rows.append({
            "odr_code": code,
            "atc_code": atc,
            "ingredient": ingredient,
            "brand_name": brand,
            "unit": unit,
            "is_antibiotic": "1" if a else "0",
            "kidney_relevant": kidney_relevant,
            "rxnorm_cui": rxnorm_cui,
            "source": source,
            "notes": notes,
        })

    with FORMULARY_CSV.open("w", encoding="utf-8", newline="") as fout:
        writer = csv.DictWriter(
            fout,
            fieldnames=[
                "odr_code", "atc_code", "ingredient", "brand_name", "unit",
                "is_antibiotic", "kidney_relevant", "rxnorm_cui", "source", "notes",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    # Gap report: DB codes not covered by formulary/abx/legacy.
    # Rows that the user already filled in (present in gap_fills) stay in the
    # file so edits aren't lost — re-running the build is safe & idempotent.
    gap_rows: list[dict[str, Any]] = []
    filled_count = 0
    for code, (cnt, sample) in sorted(db_freq.items(), key=lambda x: -x[1][0]):
        in_formulary = code in formulary or code in abx or code in legacy
        if in_formulary:
            continue
        existing = gap_fills.get(code, {})
        if existing.get("atc_code"):
            filled_count += 1
        gap_rows.append({
            "odr_code": code,
            "db_row_count": cnt,
            "sample_name": existing.get("sample_name") or sample,
            "suggested_atc": existing.get("atc_code", ""),
            "notes": existing.get("notes", ""),
        })

    with GAPS_CSV.open("w", encoding="utf-8", newline="") as fout:
        writer = csv.DictWriter(
            fout,
            fieldnames=["odr_code", "db_row_count", "sample_name", "suggested_atc", "notes"],
        )
        writer.writeheader()
        writer.writerows(gap_rows)

    return len(rows), len(gap_rows), filled_count


def main() -> int:
    print("Building unified formulary CSV...")
    formulary = load_formulary_xlsx()
    print(f"  formulary xlsx:  {len(formulary)} codes ({sum(1 for v in formulary.values() if v['atc_code'])} with ATC)")

    abx = load_abx_xlsx()
    print(f"  abx xlsx:        {len(abx)} codes")

    legacy = load_atc_drugs_csv()
    print(f"  legacy CSV:      {len(legacy)} codes")

    db_freq = get_db_odr_frequencies()
    if db_freq:
        print(f"  DB medications:  {len(db_freq)} distinct ODR_CODEs")

    gap_fills = load_gap_fills()
    if gap_fills:
        print(f"  Manual gap fills: {len(gap_fills)} codes with user-supplied ATC")

    n_rows, n_gaps, n_filled = merge_and_write(formulary, abx, legacy, db_freq, gap_fills)
    print(f"\nWrote {n_rows} rows → {FORMULARY_CSV.relative_to(REPO_ROOT)}")
    print(f"Wrote {n_gaps} gap rows ({n_filled} already filled) → {GAPS_CSV.relative_to(REPO_ROOT)}")

    # Source breakdown
    with FORMULARY_CSV.open(encoding="utf-8") as f:
        breakdown: dict[str, int] = defaultdict(int)
        atc_present = 0
        for row in csv.DictReader(f):
            breakdown[row["source"]] += 1
            if row["atc_code"]:
                atc_present += 1
    print(f"\nSource breakdown:")
    for src, n in sorted(breakdown.items()):
        print(f"  {src:<18s} {n:>5}")
    print(f"  total with ATC:    {atc_present:>5}")

    # Gap budget vs DB
    if db_freq:
        total_db_rows = sum(cnt for cnt, _ in db_freq.values())
        total_gap_rows = sum(row["db_row_count"] for row in [
            # re-read gaps
        ])
        with GAPS_CSV.open(encoding="utf-8") as f:
            gap_rows_pct = sum(int(row["db_row_count"]) for row in csv.DictReader(f))
        print(f"\nGap coverage against DB:")
        print(f"  DB rows not covered: {gap_rows_pct} / {total_db_rows} = {100*gap_rows_pct/total_db_rows:.1f}%")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
